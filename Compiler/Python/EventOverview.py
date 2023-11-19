from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import Series, Reference, ScatterChart, BarChart
from openpyxl.workbook import Workbook
import pandas as pd
from util import makeCall
from io import StringIO
from openpyxl.chart.shapes import GraphicalProperties

year = 2023
eventKey = "2023mrcmp"
districtKey = "2023fma"
dataframe = pd.DataFrame()

eventTeams = pd.read_json(StringIO(makeCall("TBA", f"event/{eventKey}/teams/simple").text))
dataframe.insert(0, "key", eventTeams.get("key"))
dataframe.insert(1, "Number", eventTeams.get("team_number"))
dataframe.insert(2, "Name", eventTeams.get("nickname"))

teams = dataframe.get("Number").to_list()
columnsToAdd = []
counter = 0

for team in teams:
    print(counter)
    counter += 1
    response = makeCall("Statbotics", f"team_year/{team}/{year}").text
    
    epaData = pd.read_json(StringIO(f"[{response}]"))
    columnsToAdd.append([round(epaData["epa_end"][0], 3), round(epaData["teleop_epa_end"][0], 3), round(epaData["auto_epa_end"][0], 3), round(epaData["endgame_epa_end"][0], 3), round(epaData["rp_1_epa_end"][0] + epaData["rp_1_epa_end"][0], 3), round(epaData["winrate"][0], 3), round(epaData["district_epa_rank"][0], 3)])
    

dfToAdd = pd.DataFrame(columnsToAdd)
dataframe.insert(3, "EPA", dfToAdd.get(0))
dataframe.insert(4, "Tele EPA", dfToAdd.get(1))
dataframe.insert(5, "Auto EPA", dfToAdd.get(2))
dataframe.insert(6, "End EPA", dfToAdd.get(3))
dataframe.insert(7, "RP EPA", dfToAdd.get(4))
dataframe.insert(8, "Winrate", dfToAdd.get(6))


districtData = makeCall("TBA", f"district/{districtKey}/rankings").text




dataframe = dataframe.drop(labels="key", axis='columns')
dataframe = dataframe.sort_values("EPA")
# dataframe.set_index("Number", inplace=True)

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(dataframe, index=False, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

# chart = ScatterChart(scatterStyle="marker")
chart = BarChart()
chart.title = "Winrate vs. EPA"
chart.y_axis.title = 'EPA'
chart.x_axis.title = 'Total'
chart.legend = None
chart.style = 10
chart.grouping = 'stacked'
chart.type = "bar"
chart.overlap = 100
chart.graphical_properties = GraphicalProperties()
chart.graphical_properties.line.noFill = True
chart.graphical_properties.line.prstDash = None

cats = Reference(ws, min_col=1, min_row=2, max_row=61)
chart.set_categories(cats)

t_epa = Reference(ws, min_col=4, min_row=1, max_row=61)
e_epa = Reference(ws, min_col=5, min_row=1, max_row=61)
a_epa = Reference(ws, min_col=6, min_row=1, max_row=61)
chart.add_data(t_epa)
chart.add_data(e_epa)
chart.add_data(a_epa)


ws.add_chart(chart, "J3")

wb.save("(Pre)ScoutingDataCompiler/Compiler/Output/pandas_openpyxl.xlsx")

# with pd.ExcelWriter("Compiler/Output/EventOverview.xlsx") as writer:  # doctest: +SKIP
#     dataframe.to_excel(writer, sheet_name="CompOverview")


