from openpyxl import Workbook

from openpyxl.chart import (
    DoughnutChart,
    Reference,
    Series,
    BarChart
)
from openpyxl.chart.series import DataPoint
import openpyxl


def make_gauge_chart(workbook: Workbook, percent: float = 0.5, sweep: int = 180, title: str = "Gauge Chart", sections: list = [["f54842", 1]]):
    
    
    total = 360
    ang = (( 0 - sweep/2 + sweep * percent -1) + 360) % 360
    white = total - sweep
    
    bg1 = round((sweep - 2) * percent, 3) 
    bg2 = round((sweep - 2) * (1- percent), 3)

    bg1_left = bg1
    bg1_list = []
    bg2_list = []
    for i, section in enumerate(sections):
        print(bg1_left)
        if bg1_left - (section[1]*(sweep)) >= 0:
            bg1_list.append([section[0], section[1]*sweep])
            
            bg1_left -= round(section[1]*(sweep), 3)
            
            
        else:
            
            bg1_list.append([section[0], bg1_left])
            bg2_list.append([section[0], (section[1]*(sweep-2))-bg1_left])
            bg2_list += [[next_section[0], next_section[1] * (sweep-2)] for next_section in sections[i+1:len(sections)]]
            break
    

    section_data = [
        [f"{title}: {percent * 100}%"],
        ["Total", total],
        ['000000', 2]] + bg2_list + [['FFFFFF', white]] + bg1_list + [[],
        ["Total", total],
        ['000000', 2],
        ["bg_two", total - 2],
        []
    ]
        
    # print(section_data)

    if "chart_util" in workbook.sheetnames:
        ws = workbook["chart_util"]
    else:
        workbook.create_sheet("chart_util")
        ws = workbook["chart_util"]

    for row in section_data:
        ws.append(row)
    
    
    startRow = 0
    for row in ws.iter_rows():
        startRow += 1
        if row[0].internal_value == f"{title}: {percent * 100}%": break
    
    startRow += 1
    

    

    chart = DoughnutChart(firstSliceAng=ang)
    chart.title = title
    chart.legend = None
    chart.width = 10

    data_two = Reference(ws, min_col=2, min_row=startRow+4+len(bg1_list)+len(bg2_list), max_row=startRow+6+len(bg1_list)+len(bg2_list))
    chart.add_data(data_two, titles_from_data=True)
    slices_two = [DataPoint(idx=i) for i in range(2)]
    needle_inner, white_inner = slices_two
    chart.series[0].data_points = slices_two
    white_inner.graphicalProperties.solidFill = "FFFFFF"#"441bbf"
    needle_inner.graphicalProperties.solidFill = "000000"
    


    data = Reference(ws, min_col=2, min_row=startRow, max_row=startRow+2+len(bg1_list)+len(bg2_list))
    chart.add_data(data, titles_from_data=True)
    slices = [DataPoint(idx=i) for i in range(2+len(bg1_list)+len(bg2_list))]
    
    for i, chart_slice in enumerate(slices):
        slices[i].graphicalProperties.solidFill = section_data[i+2][0]
    
    # needle, bg_two, white, bg_one = slices
    
    # white.graphicalProperties.solidFill = "FFFFFF"#"441bbf"
    # needle.graphicalProperties.solidFill = "000000"
    # bg_two.graphicalProperties.solidFill = "f54842"
    # bg_one.graphicalProperties.solidFill = "f54842"
    
    chart.series[1].data_points = slices
    
    return chart
    


file = Workbook()#openpyxl.load_workbook("(Pre)ScoutingDataCompiler/Compiler/Output/gaugeTesting.xlsx")

chart = make_gauge_chart(file, percent=.80, sweep=240, sections=[["ff0000", .2], ["f0eded", .5], ["05fa0d", .3]])
ws = file["chart_util"]
ws.add_chart(chart, "F5")



file.save("(Pre)ScoutingDataCompiler/Compiler/Output/gaugeTesting.xlsx")